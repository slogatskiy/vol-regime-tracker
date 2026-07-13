"""Metric 8 (page section 04) — REAL US corporate new-issue volume from SIFMA.

Parses SIFMA's official "US Corporate Bonds Statistics" workbook (Issuance tab) — the only
internally-consistent source for monthly IG/HY issuance. Download it once from
https://www.sifma.org/research/statistics/us-corporate-bonds-statistics/ ("US Corporate Bond
Statistics" link / email) and save it to data_raw/sifma_us_corporate.xlsx, then run this.

The Issuance tab (source: Refinitiv, $ billions) stacks three blocks under a
[Investment Grade | High Yield | Total] header (cols B/C/D):
  - annual rows (col A = a 4-digit year)
  - quarterly rows (col A = e.g. "1Q26")
  - monthly rows (col A = a month-end date)
We read all three. Monthly currently spans a trailing ~13 months.

    python scripts/ingest_new_issue.py
"""
import os
import re
import datetime as dt
from _common import save, today, RAW

XLSX = os.path.join(RAW, "sifma_us_corporate.xlsx")

# Documented record months (for callouts on the page).
RECORD_MONTHS = [
    {"month": "2025-09", "note": "record month; post-Sep rate-cut IG surge"},
    {"month": "2026-01", "note": "record January; +12% YoY, 6th month ever >$200bn IG"},
]


def _row(a, b, c, d):
    ig = round(b, 1) if isinstance(b, (int, float)) else None
    hy = round(c, 1) if isinstance(c, (int, float)) else None
    tot = round(d, 1) if isinstance(d, (int, float)) else None
    return ig, hy, tot


def parse_sifma_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Issuance"]
    annual, quarterly, monthly = [], [], []
    q_re = re.compile(r"^([1-4])Q(\d{2})$")
    for a, b, c, d in ((r[0], r[1], r[2], r[3]) for r in ws.iter_rows(values_only=True)):
        if a is None:
            continue
        ig, hy, tot = _row(a, b, c, d)
        if tot is None:
            continue
        if isinstance(a, dt.datetime):                      # monthly (month-end date)
            monthly.append({"month": a.strftime("%Y-%m"), "ig_bn": ig, "hy_bn": hy, "total_bn": tot})
        elif isinstance(a, (int, float)) and 2000 <= a <= 2100:   # annual
            annual.append({"year": str(int(a)), "ig_bn": ig, "hy_bn": hy, "total_bn": tot})
        elif isinstance(a, str):
            s = a.strip()
            if s.isdigit() and len(s) == 4:                 # annual as text
                annual.append({"year": s, "ig_bn": ig, "hy_bn": hy, "total_bn": tot})
            else:
                m = q_re.match(s)
                if m:                                       # quarterly e.g. 1Q26
                    q = f"20{m.group(2)}Q{m.group(1)}"
                    quarterly.append({"q": q, "ig_bn": ig, "hy_bn": hy, "total_bn": tot})
    monthly.sort(key=lambda r: r["month"])
    quarterly.sort(key=lambda r: r["q"])
    annual.sort(key=lambda r: r["year"])
    return annual, quarterly, monthly


def main():
    if not os.path.exists(XLSX):
        raise SystemExit(
            f"Missing {XLSX}. Download SIFMA 'US Corporate Bond Statistics' xlsx and save it there.")

    annual, quarterly, monthly = parse_sifma_xlsx(XLSX)

    # monthly IG YoY where a year-ago month exists
    ig_by_month = {r["month"]: r["ig_bn"] for r in monthly}
    for r in monthly:
        y, mo = r["month"].split("-")
        prev = f"{int(y) - 1}-{mo}"
        r["ig_yoy_pct"] = (round((r["ig_bn"] / ig_by_month[prev] - 1) * 100)
                           if prev in ig_by_month and ig_by_month[prev] else None)

    last = monthly[-1] if monthly else None
    ytd26 = next((a for a in annual if a["year"] == "2026"), None)
    save("new_issue.json", {
        "last_updated": today(),
        "illustrative": False,
        "granularity": "monthly",
        "source": "SIFMA US Corporate Bond Issuance (Issuance tab, source Refinitiv)",
        "source_url": "https://www.sifma.org/research/statistics/us-corporate-bonds-statistics/",
        "unit": "Gross issuance, $ billions",
        "record_months": RECORD_MONTHS,
        "annual": annual,
        "quarterly": quarterly,
        "monthly": monthly,
        "note": "Real SIFMA monthly IG/HY gross issuance. IG boom: ~$1.27T (2023) → $1.57T (2024) "
                "→ $1.73T (2025), a record year. Jan-2026 IG $231bn and Sep-2025 IG $211bn are the "
                "standout prints behind the secondary-turnover surge; Dec-2025 shows the usual "
                "year-end lull. High new-issue activity drives secondary turnover.",
    })
    msg = f"new issue: {len(monthly)} monthly, {len(quarterly)} quarterly, {len(annual)} annual rows"
    if last:
        msg += f"; latest {last['month']} IG ${last['ig_bn']}bn"
    print(msg)


if __name__ == "__main__":
    main()
